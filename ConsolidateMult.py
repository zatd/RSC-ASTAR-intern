# -*- coding: utf-8 -*-
"""
Created on Wed Apr 17 10:11:59 2019

@author: tp-AizatKA
"""

import os
from openpyxl import load_workbook
import pandas as pd
import sys

#pip install --upgrade easygui
from easygui import enterbox, choicebox, msgbox, multchoicebox

def unmerge_file(filename, sheetname):
    
    wb = load_workbook(filename)
    ws = wb[sheetname]
        
    #df = pd.read_excel(filename, sheetname)
    
    while True:
        count = 0
        for i in ws.merged_cells.ranges:
            count += 1
            ws.unmerge_cells(str(i))
        if count == 0: break
    
    unmerged_filename = filename.split('.')[0] + ' new.xlsx'
    wb.save(unmerged_filename)
    wb.close()
    
    #delete unwanted sheets
    wb = load_workbook(unmerged_filename)
    for sheet_to_delete in reversed(wb.sheetnames):
        if sheet_to_delete != sheetname: del wb[sheet_to_delete]
    wb.save(unmerged_filename)
    return unmerged_filename
    
def clean_compile (unmerged_filename, sheetname):
    
    df = pd.read_excel(unmerged_filename, sheetname, skiprows = 5)
    df = df.drop([0], axis=0)
    df = df.drop(['Unnamed: 0'], axis=1)
    #get list to be deleted
    del_list = []
    for col in df.columns:
        if col in ['Unnamed: 1', 'Unnamed: 2']:
            for row in range(0, len(df)):
                #check for nan
                if pd.isna(df.iloc[row][col]):
                    #.at and .iloc is of different indexing rules
                    df.at[row+1, col] = df.iloc[row-1][col]
        if col == 'Brand':
            for row in range(0, len(df)):
                if not pd.isna(df.iloc[row][col]):
                    if df.iloc[row][col].lower() not in ['no offer', 'no bid']:
                        continue
                del_list.append(row)
    #delete unwanted rows
    for index in reversed(del_list):
        df = df.drop([index+1], axis=0)
       
    #re-index entries
    df = df.reset_index(drop=True)
    
    #delete last row
    df = df.drop([len(df)-1], axis=0)
        
    return df

def combine_lists(data_file):
    
    df = pd.concat(pd.read_excel(data_file, sheet_name=None, skiprows=0), sort=False)
    df = df.reset_index(level=1, drop=True).rename_axis('Sheet Names').reset_index()
    
    book = load_workbook(data_file)
    writer = pd.ExcelWriter(data_file, engine = 'openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name = 'All')
    writer.save()
    writer.close()

    
def main():
    
    #will check for all files from script directory onwards
    print (os.getcwd(), ":") #get dir address for current file
    files_dir = list()
    for root, dirs, files in os.walk('.'):
        for file in files:
            if '.xlsx' in file:
                files_dir.append(file)
                print ("\t", file)
                
    print ("")
    #dialog box to choose file
    msg ="Please choose file to consolidate or Press <Cancel> to exit."
    title = "Directory List of .xlsx Files"
    choice_list = {}
    choice_list = multchoicebox(msg, title, files_dir)

    if choice_list is None:
        sys.exit("No file was selected..")
        
    sheetname = 'Part 6 Price Schedule'
    data_file = 'output.xlsx'
    writer = pd.ExcelWriter(data_file)
    
    for filename in choice_list:
        unmerged_filename = unmerge_file(filename, sheetname)
        df = clean_compile (unmerged_filename, sheetname)
        os.remove(unmerged_filename)
        df.to_excel(writer, sheet_name = filename.split('.')[0])
    writer.save()
    writer.close()
    
    combine_lists(data_file)
    
    print ("Compilation complete!")
if __name__ == '__main__':
    main()
