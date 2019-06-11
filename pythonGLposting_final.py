# =============================================================================
# Objective: to generate report in the given format for GL posting
# - ensure files are in .xlsx or .csv
# - ensure files have 22 columns: 'A' to 'V'
# - ensure files only consist of raw data: no extra rows added below
# - extraction of data is by keyword extractname which is set to 'Bill-to (non-A*STAR)/GL Code (A*STAR)' in main
# - any blank cells in this column will be converted to '0'
# - ensure files to be processed are in the same folder as this script
# - ensure no other .xlsx or .csv file is in the same folder or its subfolder
# - to change the name format of the generated file, use keyword finalfilename
# =============================================================================

import os
import pandas as pd
import openpyxl as opx

from easygui import msgbox, enterbox
#need installation of easygui on anaconda prompt:[pip install easygui]

import sys

from datetime import datetime

#converts .csv files to .xlsx files for compatibality
def convert_to_xlsx(init_filename):
    import csv
    from openpyxl import Workbook

    
    f = open(init_filename)
    csv.register_dialect('comma', delimiter=',')
    reader = csv.reader(f, dialect='comma')
    
    wb = Workbook()
    dest_filename = init_filename.replace('.csv', '.xlsx')
    
    ws = wb.worksheets[0]
    ws.title = "Sheet 1"
    
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):

            ws.cell(row_index + 1, column_index + 1).value = cell
    
    wb.save(filename = dest_filename)
    return dest_filename

#clean whole file from special character '|'
def clean(file):
    wb = opx.load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in range(1, ws.max_column):
            for row in range(1, ws.max_row + 1):
                if type(ws.cell(row, col).value) is str:
                    # .replace('changefrom', 'changeto') and .strip() removes leading/trailing spaces
                    ws.cell(row, col).value = ws.cell(row, col).value.replace('|', '').strip()
                    ws.cell(row, col).value = ws.cell(row, col).value.replace('ï»¿', '').strip()
                    if ws.cell(row, col).value is not None and ws.cell(row, col).value.isdigit(): ws.cell(row, col).value = int(ws.cell(row, col).value)
                #assign 0 for blanks under 'Bill to....' column
                if col == 2 and ws.cell(row, col).value is None: ws.cell(row, col).value = 0
    wb.save(file)
 
#get description for blue summary table
def get_des(file):
    wb = opx.load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in range(1, ws.max_column):
            if ws.cell(1, col).value == 'Description':
                description = ws.cell(2, col).value.rsplit('-',1)[0]
    wb.save(file)
    return description
    
#separate external and internal
def extract_ext(file_raw, new_file, extractkey):
    #open excel file as df by pandas
    df = pd.read_excel(file_raw)
    writer = pd.ExcelWriter(new_file + '.xlsx')
    
    sheet_name_list = list()
    for types in df.filter([extractkey])[extractkey].unique():
        if isinstance(types, str): sheet_name_list.append('Other')
        elif types < 1000000: sheet_name_list.append('Internal')
        elif types >= 1000000: sheet_name_list.append('External')
        
    for ltype in reversed(sheet_name_list):
        list_type = list()
        
        for index_i, i in enumerate(df.filter([extractkey])[extractkey]):
            if ltype == 'External': condition = isinstance(i, int) and i >= 1000000
            elif ltype == 'Internal': condition = isinstance(i, int) and i < 1000000
            elif ltype == 'Other': condition = isinstance(i, str) and (i[5:] == 'IBN' or i[5:] == 'GIS')
            
            if condition == True: list_type.append(index_i)
            
        #insert list to dataframe
        dftype = df.loc[list_type]
        
        dftype.sort_values('A*STAR RI Cost Centre', inplace = True)
        #write to file
        dftype.to_excel(writer, sheet_name = ltype)
        
    writer.save()
    writer.close()
    return sheet_name_list
                
def gl_posting(sortedfile, finalfilename, sheetname, description, static_data):
    #open excel file as df by pandas
    df_read = pd.read_excel(sortedfile + '.xlsx', sheet_name = sheetname)
    df_ri = []
    
    #remove 'Billing Address', 'SAP material code', 'PO Number (non-A*STAR only)', 'Amount to bill', 'GST'
    columns_to_delete = ['Billing Address', 'SAP material code', 'PO Number (non-A*STAR only)', 'Amount to bill', 'GST']
    for columnname in columns_to_delete:
        del df_read[columnname]
    
    #add RI column
    for no_of_entries, i in enumerate(df_read.filter(['Account number'])['Account number']):
        df_ri.append(i.split('_')[0].upper())
    df_read.insert(2, 'RI', df_ri)
    df_read.sort_values(['RI', 'A*STAR RI Cost Centre'], inplace = True)
    
    #get today's date MMMYY
    current_year = datetime.now().strftime('%y')
    current_month_text = datetime.now().strftime('%b')
    mmmyy = current_month_text.upper() + current_year
    
    #create filename
    if sheetname == 'Internal':
        finalfilename += '_GL code.xlsx'
    else: finalfilename += ' - ' + sheetname + '.xlsx'
    
    #write to new excel file
    with pd.ExcelWriter(finalfilename) as writer:  # doctest: +SKIP
        df_read.to_excel(writer, sheet_name='Journal Doc_PPMS')
    
    #open file using openpyxl for further editing
    wb = opx.load_workbook(finalfilename)
    ws = wb['Journal Doc_PPMS']
    
    new_line = []
    #separate RIs
    cell_check = ws['D2'].value
    index_RI = 2
    while True:
        index_RI += 1
        if (ws['D%d' % (index_RI)].value) != cell_check and ws['D%d' % (index_RI)].value is not None:
            cell_check = ws['D%d' % (index_RI)].value
            ws.insert_rows(index_RI, amount = 1)
            new_line.append(index_RI)
            #copy column header for new group
            for i in range(1, 20):
                cellref = ws.cell(index_RI, i)
                cellref.value = ws.cell(1, i).value
                header_font(cellref) #all except top most header is formatted here
        #if next five lines are blank, stop loop
        end = False
        for i in range(0,4):
            if ws['D%d' % (index_RI + i)].value is not None: break
            else: end = True
        if end == True: break
    
    #run through all rows, convert column O to proper date and column Q numbers to currency and borders for body
    for row_number in range(1, ws.max_row + 1):
        if ws.cell(row_number, 1).value != 'Account number' and ws.cell(row_number, 1).value is not None:
            currency_font(ws.cell(row_number, 18))
            dateformat_font(ws.cell(row_number, 16))
            for col_number in range(1, ws.max_column + 1):
                body_font(ws.cell(row_number, col_number))
    for i in reversed(new_line):
        ws.insert_rows(i, amount = 2)
    total_amount_list = list()
    #get totals for Total amount
    for column, column_names in enumerate(ws['1'], 1):
        header_font(ws.cell(1, column))
        grand_total = 0
        if column_names.value == 'Total amount': #by title of column U
            for row in range(1, ws.max_row + 2):
                if ws.cell(row, column).value == 'Total amount':
                    subtotal = 0
                elif ws.cell(row, column).value is not None: 
                    subtotal += ws.cell(row, column).value
                elif subtotal != None: 
                    ws.cell(row, column).value = subtotal
                    ws.cell(row, column - 1).value = "Total Amount"
                    totals_font(ws.cell(row, column))
                    totals_font(ws.cell(row, column - 1))
                    grand_total += subtotal
                    total_amount_list.append(subtotal)
                    subtotal = None
            rmax = ws.max_row + 2
            ws.cell(rmax, column).value = grand_total
            ws.cell(rmax, column - 1).value = "Grand Total"
            totals_font(ws.cell(ws.max_row, column))
            totals_font(ws.cell(ws.max_row, column - 1))

    #insert document head
    ws.delete_cols(1, amount = 1)
    ws.insert_rows(0, amount = 7)
    ws['A2'] = "Reference"
    ws['A4'] = "Doc Header Text"
    ws['C2'] = "RSC PPMS - " + convertto_mmmyy(ws.cell(9, 11).value)
    ws['C4'] = "RSC PPMS BILLING - " + convertto_mmmyy(ws.cell(9, 11).value)
    ws['A7'] = "For AP Entries"
    
    #insert summary table (blue)
    summary_header = ['Item', 'PK (DR/CR)', 'GL code', '', '',
                      'Amount S$ (DR)', 'Amount S$ (CR)', 'Text*', 
                      'BA ( business area)', 'Cost Centre', 'Fund', 
                      'WBS element', 'Funds Center', 'Commt item']
    ws.insert_rows(7, amount = 3)
    for column, column_name in enumerate(summary_header, 1):
        ws.cell(7, column).value = column_name
        summaryhead_font(ws.cell(7, column))

    ri_details = list()
    index = 0
    extraction = False
    for row_number in range(2, ws.max_row):
        if ws.cell(row_number-1, 1).value == 'Account number':
            extraction = True
            index += 1
        elif ws.cell(row_number, 1).value is None:
            extraction = False
        #extract from row_number + 1
        if extraction == True:
            #ws.cell(row_number, 3).value is RI
            a=[ws.cell(row_number, 3).value]
            a.extend(['CR', '100010', '', '', ''])
            #ws.cell(row_number, 17).value is Amount CR
            a.extend([ws.cell(row_number, 17).value])
            a.extend(['RSC-' + ws.cell(row_number, 3).value.upper() + ' ' + convertto_mmmyy(ws.cell(row_number, 11).value)+ " " + description])
            a.extend(static_data)
            a.extend([ws.cell(row_number, 9).value, ws.cell(row_number, 2).value])
            #ws.cell(row_number, 9).value is the cost centre and ws.cell(row_number, 2).value is the Bill-to
            ri_details.append(a)
                               
    
    #collate those CR of same cost centre in list
    buffer = None
    del_list = list()
    for index, content in enumerate(ri_details):
        if buffer is None or buffer != ri_details[index][-2]: 
            buffer = ri_details[index][-2]
            amount = ri_details[index][6]
            counter = 1
        else:
            amount += ri_details[index][6]
            ri_details[index-counter][6] = amount
            del_list.append(index)
            counter += 1
    #delete collapsed rows
    for index in reversed(del_list):
        del ri_details[index]
    
    #insert DR for each RI in list    
    buffer = None
    insert_index = None      
    subtotal_list = list()
    for index, entry in enumerate(ri_details):
        if buffer is None or buffer != ri_details[index][0]:
            if insert_index is not None:
                ri_details.insert(insert_index, (buffer, 'DR', str(ri_details[insert_index][-1]), '', '', amount, '', ri_details[insert_index][7], static_data[0], '' , '' , '', static_data[4], ''))
                subtotal_list.append(amount)
                index += 1
                amount = 0
            else: amount = ri_details[index][6]
            insert_index = index
            buffer = ri_details[index][0]
        else: amount += ri_details[index][6]
    ri_details.insert(insert_index, (buffer, 'DR', str(ri_details[insert_index][-1]), '', '', amount, '', ri_details[index][7], static_data[0], '' , '' , '', static_data[4], ''))
    subtotal_list.append(amount)

    #use this to collate total for amount columns
    dr_total = 0
    cr_total = 0
    #inserting the summary table
    for summary_row in range(1, len(ri_details) + 1):
        ws.insert_rows(summary_row + 7, amount = 1)
        if type(ri_details[summary_row-1][5]) != str:   
            dr_total += ri_details[summary_row-1][5]
        elif type(ri_details[summary_row-1][6]) != str: cr_total += ri_details[summary_row-1][6]
        
        for summary_col in range(1, len(ri_details[0]) + 1):
            # + 7 to insert data under summary header
            ws.cell(summary_row + 7, summary_col).value = ri_details[summary_row-1][summary_col-1]
            summary_font(ws.cell(summary_row + 7, summary_col))
            if summary_col == 5 or 6:
                currency_font(ws.cell(summary_row + 7, summary_col))
    
    #sum up total amount of summary table
    ws.insert_rows(summary_row + 7 + 1, amount = 1)
    ws.cell(summary_row + 7 + 1, 5).value = 'Total Amount'
    ws.cell(summary_row + 7 + 1, 6).value = dr_total
    ws.cell(summary_row + 7 + 1, 7).value = cr_total
    totals_font(ws.cell(summary_row + 7 + 1, 5))
    totals_font(ws.cell(summary_row + 7 + 1, 6))
    totals_font(ws.cell(summary_row + 7 + 1, 7))
    
    #perform amounts checks
    #column 16 is column Q
    index_subtotal_list = 0
    no_error_bool = True
    for row in range(1, ws.max_row + 1):
        amount_check = False
        amount_check_symbol = u'\u2717'
        if ws.cell(row, 16).value == 'Grand Total':
            if round(dr_total, 2) == round(cr_total, 2) == round(ws.cell(row, 16 + 1).value, 2):
                amount_check = True
                amount_check_symbol = u'\u2713'
            else: no_error_bool = False
            ws.cell(row, 16 + 2).value = amount_check
            print ("Grand Total check:", "%.2f" % dr_total, "%.2f" % cr_total, "%.2f" % ws.cell(row, 16 + 1).value, amount_check_symbol)
        if ws.cell(row, 16).value == 'Total Amount':
            if round(subtotal_list[index_subtotal_list], 2) == round(ws.cell(row, 16 + 1).value, 2):
                amount_check = True
                amount_check_symbol = u'\u2713'
            else: no_error_bool = False
            index_subtotal_list += 1
            ws.cell(row, 16 + 2).value = amount_check
            print ("Total Amount check " + str(index_subtotal_list) + ": ",
                   "%.2f" % subtotal_list[index_subtotal_list-1], "%.2f" % ws.cell(row, 16 + 1).value,
                   amount_check_symbol)
            
    #pop up alert if error exists
    if no_error_bool == False:
        # A nice welcome message
        ret_val = msgbox("Error! Please check amount in: \n" + finalfilename, "Alert!")
        if ret_val is None: # User closed msgbox
            sys.exit(0)
        
    #set column widths
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 32
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 13
    ws.column_dimensions['Q'].width = 23
    ws.column_dimensions['R'].width = 23
    
    wb.save(finalfilename)
    
    #print completion on console
    print (finalfilename, "done!\n")
    return finalfilename

def add_separation (completedfile):
    #open excel file as df by openpyxl
    wb = opx.load_workbook(completedfile)
    ws = wb['Journal Doc_PPMS']
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == 'CR' and ws.cell(row + 1, 2).value == 'DR': 
            ws.insert_rows(row + 1, amount = 1)
            for col in range(1, 15): #15 coloumns in summary table
                summary_font(ws.cell(row + 1, col))
        if ws.cell(row, 1).value == 'For AP Entries': break
    wb.save(completedfile)
        
#formatting cells section~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

#colors codes can be found here: http://dmcritchie.mvps.org/excel/colors.htm

def header_font (cell):
    cell.alignment = Alignment(horizontal='left')
    cell.font = Font(bold = False)
    cell.fill = PatternFill(fgColor = 'FFFF00', fill_type = 'solid')
    cell.border = Border(left = Side(style = 'thin'), 
                         right = Side(style = 'thin'), 
                         top = Side(style = 'thin'), 
                         bottom = Side(style = 'thin'))
def body_font (cell):
    cell.alignment = Alignment(horizontal='right')
    cell.font = Font(bold = False)
    cell.border = Border(left = Side(style = 'thin'), 
                         right = Side(style = 'thin'), 
                         top = Side(style = 'thin'), 
                         bottom = Side(style = 'thin'))
def totals_font (cell):
    cell.font = Font(bold = True)
    cell.border = Border(bottom = Side(style = 'double'))
    cell.number_format = '#,##0.00'
def summaryhead_font (cell):
    cell.fill = PatternFill(fgColor = 'CCCCFF', fill_type = 'solid')
    cell.border = Border(left = Side(style = 'thin'), 
                         right = Side(style = 'thin'), 
                         top = Side(style = 'thin'), 
                         bottom = Side(style = 'thin'))
def summary_font (cell):
    cell.fill = PatternFill(fgColor = '99CCFF', fill_type = 'solid')
    cell.border = Border(left = Side(style = 'thin'), 
                         right = Side(style = 'thin'), 
                         top = Side(style = 'thin'), 
                         bottom = Side(style = 'thin'))
def currency_font (cell):
    cell.number_format = '#,##0.00'
def dateformat_font (cell):
    cell.number_format = 'dd/mm/yyyy'
def convertto_mmmyy (string):
    from datetime import datetime
    # you could also import date instead of datetime and use that.
    date = datetime(year=int(string[-8:-4]), month=int(string[-4:-2]), day=int(string[-2:]))
    return date.strftime("%b%y").upper()

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

def main():
    tempfilename = 'Tempfile'
    extractname = 'Bill-to (non-A*STAR)/GL Code (A*STAR)'
    
    #will check for all files from specified directory onwards
    cwd = os.getcwd()
    change_dir = enterbox(msg="Please enter the directory to process: ",title="GL Posting",default=cwd)

    # trying to insert directory 
    try:
        os.chdir(change_dir) 
        print("Change directory to...")
        print(os.getcwd() + ':') 
    # Catching the exception     
    except: 
        print("Something wrong with specified directory. Exception- ", sys.exc_info())
    
    dir_list = list()
    for root, dirs, files in os.walk('.'):
        for file in files:
            if '.xlsx' in file or '.csv' in file :
                # to prevent running own generated file
                if '[PY' not in file :
                    dir_list.append(file)
                    print ("\t", file)
    print ("")
    
    for filename in dir_list:
        temp_created = False
        if filename[-4:]=='.csv': 
            filename = convert_to_xlsx(filename)
            temp_created = True
        
        #get today's date YYYYMMDD
        current_year_full = datetime.now().strftime('%Y')
        current_month = datetime.now().strftime('%m')
        current_day = datetime.now().strftime('%d')
        yyyymmdd = current_year_full + current_month + current_day
        
        #check if file is of same column length as expected (low level checking)
        df = pd.read_excel(filename)
        if len(df.columns) == 22:
            
            print ("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            print (filename)
            print ("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            clean(filename)
            description = get_des(filename)
            
            """
            SET YOUR VARIABLES HERE
            """
            finalfilename = "[PY " + yyyymmdd + "] " + filename.replace('.xlsx', '')
            static_data = ['SC1', 'X00020' , 'RSC19' , '', 'X00001', '100000']
            """
            SET YOUR VARIABLES HERE
            """
            
            sheet_name_list = extract_ext(filename, tempfilename, extractname)
            for sheet_name in reversed(sheet_name_list):
                if sheet_name != 'External':
                    completedfile = gl_posting(tempfilename, finalfilename, sheet_name, description, static_data)
                    add_separation (completedfile)
            os.remove(tempfilename + '.xlsx')
            if temp_created == True:
                os.remove(filename)
            #print ("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        else: print (u'\u2717', "Error! File '" + filename + "' is not of reasonable format for this operation")
        
if __name__ == '__main__':
    main()