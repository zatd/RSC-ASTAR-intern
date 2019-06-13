# -*- coding: utf-8 -*-
"""
Created on Fri Jun  7 14:55:10 2019

@author: tp-AizatKA
"""

#Customer List.xlsx file name MUST strictly be "Customer List.xlsx"
#Transfer of Revenue to RIs - FY19 file name MUST contain "Transfer of Revenue to RIs"

import os
import numpy
import calendar
import pandas as pd
import openpyxl as opx
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

#converts .csv files to .xlsx files for compatibality
def convert_to_xlsx(init_filename):
    import csv
    from openpyxl import Workbook

    
    f = open(init_filename)
    csv.register_dialect('comma', delimiter=',')
    reader = csv.reader(f, dialect='comma')
    
    wb = Workbook()
    dest_filename = init_filename.replace('.csv', '.xlsx')
    
    ws = wb.worksheets[0]
    ws.title = "Sheet 1"
    
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):

            ws.cell(row_index + 1, column_index + 1).value = cell
    
    wb.save(filename = dest_filename)
    return dest_filename

#clean whole file from special character '|' and 'ï»¿'
def clean(file):
    wb = opx.load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in range(1, ws.max_column):
            for row in range(1, ws.max_row + 1):
                if type(ws.cell(row, col).value) is str:
                    # .replace('changefrom', 'changeto') and .strip() removes leading/trailing spaces
                    ws.cell(row, col).value = ws.cell(row, col).value.replace('|', '').strip()
                    ws.cell(row, col).value = ws.cell(row, col).value.replace('ï»¿', '').strip()
                    if ws.cell(row, col).value is not None and ws.cell(row, col).value.isdigit(): ws.cell(row, col).value = int(ws.cell(row, col).value)
                #assign 0 for blanks under 'Bill to....' column
                if col == 2 and ws.cell(row, col).value is None: ws.cell(row, col).value = 0
    wb.save(file)

def extract_pis(filename):
    wb = opx.load_workbook(filename)
    ws = wb['Completed']
    for search_col in range(1, ws.max_column+1):
        if ws.cell(1, search_col).value == 'RI PIC':
            prev_value = None
            df = []
            for row in range(2, ws.max_row+1):
                sub_df = []
                if df!=[] and prev_value != ws.cell(row, search_col).value:
                    generate_report(df, search_col, filename)
                    df = []
                for col in range(1, ws.max_column+1):
                    sub_df.append(ws.cell(row, col).value)
                df.append(sub_df)
                prev_value = ws.cell(row, search_col).value
            generate_report(df, search_col, filename)
            
def generate_report(df, search_col, main_filename):
#    for row in range(0, len(df)):
#        print (df[row][search_col-1].split('\n')[0])
    
    filename = 'Customer List.xlsx'
    wb = opx.load_workbook(filename)
    for sheet in wb.sheetnames:
        if sheet in df[len(df)-1][search_col-1]: 
            
            writer = pd.ExcelWriter(sheet + ' Transfer Report.xlsx')
            
            df_columns = pd.read_excel(filename, sheet_name=sheet)
            extractkey = 'Category'
            column_titles = df_columns.filter([extractkey])[extractkey].unique()
            
            #insert column headers
            part1 = numpy.transpose(["FY19", "Amount billed to AMP users (excluding GST)", "Amount to be retained by RSC (7%)", "Amount due to be transferred to AMP-SRIS (Column B - Column C)"])
            part2 = numpy.transpose(["Monthly amount transferred to AMP-SRIS", "Monthly outstanding amount to be transferred to AMP-SRIS", "Cumulative outstanding amount to be transferred to AMP-SRIS"])
            final_columns = numpy.concatenate((part1,column_titles,part2))
            
            #insert first row
            contents = numpy.full((1,len(final_columns)),"", dtype=object)
            contents[0,0] = "Balance carry forward from FY18"
            
            #insert fy months rows
            for i in range(1,12+1):
                financial_year_month = numpy.full((1,len(final_columns)),"", dtype=object)
                financial_year_month[0,0] = calendar.month_name[(i+2)%12+1]
                financial_year_month[0,2] = "=IFERROR(B%d-D%d,)" %(i+2,i+2)
                financial_year_month[0,3] = "=IFERROR(SUM(E%d:%s%d),)" %(i+2,chr(68+len(column_titles)),i+2)
                financial_year_month[0,len(final_columns)-2] = "=IFERROR(D%d-%s%d,)" %(i+2,chr(69+len(column_titles)),i+2)
                financial_year_month[0,len(final_columns)-1] = "=IFERROR(L%d+K%d,)" %(i+1,i+2)
                contents = numpy.append(contents, financial_year_month, axis=0)
            
            #insert totals row
            totals = numpy.full((1,len(final_columns)),"", dtype=object)
            totals[0,0] = "TOTALS for FY19"
            for i in range(1, len(totals[0])-1):
                totals[0,i] = "=IFERROR(SUM(%s3:%s14),)" %(chr(i+65), chr(i+65))
            contents = numpy.append(contents, totals, axis=0)
            
            #insert two "blank" row
            blank = numpy.full((1,len(final_columns)),"", dtype=object)
            blank[0,4] = "=SUM(E15:%s15)-D15" %(chr(68+len(column_titles)))
            blank[0,4+len(column_titles)] = "=IFERROR(SUM(%s15:%s15)=D15,)" %(chr(69+len(column_titles)), chr(70+len(column_titles)))
            contents = numpy.append(contents, blank, axis=0)
            contents = numpy.append(contents, numpy.full((1,len(final_columns)),"", dtype=object), axis=0)
            
            #insert summary block header
            summary_head = numpy.full((1,len(final_columns)),"", dtype=object)
            summary_head[0,0] = "Sources of Income"
            summary_head[0,1] = "FY19 to date"
            summary_head[0,2] = "Monthly average"
            contents = numpy.append(contents, summary_head, axis=0)
            
            #trim cutomer list data
            for column_title_index, column_title in enumerate(column_titles):
                column_titles[column_title_index] = column_titles[column_title_index].replace('-derived  revenue (after RSC 7%)', ' ')
            
            #insert summary block
            for i in range(0, len(column_titles)):
                summary = numpy.full((1,len(final_columns)),"", dtype=object)
                summary[0,0] = column_titles[i]
                summary[0,1] = "=IFERROR(%s15,)" %(chr(i+69))
                summary[0,2] = "=IFERROR(B%d/COUNT(%s$3:%s$14),)" %(i+19, chr(i+69), chr(i+69))
                contents = numpy.append(contents, summary, axis=0)
            summary_total = numpy.full((1,len(final_columns)),"", dtype=object)
            summary_total[0,0] = "Total"
            summary_total[0,1] = "=IFERROR(SUM(B19:B%d),)" %(18+len(column_titles))
            contents = numpy.append(contents, summary_total, axis=0)
            
            #insert "blank" row
            blank = numpy.full((1,len(final_columns)),"", dtype=object)
            blank[0,1] = "=B24-D15"
            contents = numpy.append(contents, blank, axis=0)
            
            df_contents = pd.DataFrame(contents, columns = final_columns)
            df_contents.to_excel(writer, index=False)
            
            writer.save()
            writer.close()
            
            format_report(sheet + ' Transfer Report.xlsx')
            
            fill_report(sheet + ' Transfer Report.xlsx', main_filename)
            
            
def format_report(filename):
    
    wb = opx.load_workbook(filename)
    ws = wb['Sheet1']
    
    #set wrap text
    for i in range(1,ws.max_row+1):
        for j in range(1,ws.max_column+1):
            ws.cell(i,j).alignment = Alignment(wrap_text=True)
    
    #set column widths
    ws.column_dimensions['A'].width = 20
    for index in range(1,4):
        ws.column_dimensions['%s' %(chr(65+index))].width = 17
    for index in range(1,ws.max_column-6):
        ws.column_dimensions['%s' %(chr(65+3+index))].width = 20
    for index in range(ws.max_column-3,ws.max_column):
        ws.column_dimensions['%s' %(chr(65+index))].width = 17
    
    #set headers fill color
    for index in range(1,ws.max_column+1):
        ws.cell(1,index).fill = PatternFill(fgColor = 'CCCCCC', fill_type = 'solid')
        
    #set column D fill color
    for index in range(1,15):
        ws.cell(index,4).fill = PatternFill(fgColor = '44DD77', fill_type = 'solid')
        
    #set added columns fill color
    for index in range(1,15):
        for repeat in range(1,ws.max_column-6):
            ws.cell(index,4+repeat).fill = PatternFill(fgColor = 'FFFF00', fill_type = 'solid')
        
    for index in range(1,4):
        ws.cell(18,index).fill = PatternFill(fgColor = 'C0C0C0', fill_type = 'solid')
    
    #set second row fill color
    for index in range(1,ws.max_column+1):
        ws.cell(2,index).fill = PatternFill(fgColor = '99CCFF', fill_type = 'solid')
        
    #set red fill color
    ws.cell(15,2).fill = PatternFill(fgColor = 'FF0000', fill_type = 'solid')
    ws.cell(15,3).fill = PatternFill(fgColor = 'FF0000', fill_type = 'solid')
    ws.cell(15,4).fill = PatternFill(fgColor = 'FF0000', fill_type = 'solid')
    ws.cell(ws.max_row-1,2).fill = PatternFill(fgColor = 'FF0000', fill_type = 'solid')
    
    #set colorless font color
    for j in range(1,ws.max_column+1):
        ws.cell(16,j).font = Font(color='CCCCFF')
        ws.cell(ws.max_row,2).font = Font(color='CCCCFF')
        
    #add border
    for j in range(1,ws.max_column+1):
        for i in range(1,15+1):
            ws.cell(i,j).border = Border(left = Side(style = 'thin'), 
                                            right = Side(style = 'thin'),
                                            top = Side(style = 'thin'), 
                                            bottom = Side(style = 'thin'))
    for j in range(1,3+1):
        for i in range(18,ws.max_row+1-1):
            ws.cell(i,j).border = Border(left = Side(style = 'thin'), 
                                            right = Side(style = 'thin'),
                                            top = Side(style = 'thin'), 
                                            bottom = Side(style = 'thin'))
    
    #bold fontstyle
    for index in range(1,ws.max_column+1):
        ws.cell(1,index).font = Font(bold = True)
        ws.cell(2,index).font = Font(bold = True)
        ws.cell(15,index).font = Font(bold = True)
        ws.cell(18,index).font = Font(bold = True)
        ws.cell(ws.max_row-1,index).font = Font(bold = True)
            
    wb.save(filename)
    wb.close()

def fill_report(filename, main_filename):
    
    ri_pic = filename.replace(' Transfer Report.xlsx','')
#    print ()
#    print (ri_pic)
#    print ()
    df_1 = pd.read_excel('Customer List.xlsx', sheet_name = ri_pic)
    df_2 = pd.read_excel(main_filename, sheet_name = 'Completed')
    df_3 = pd.read_excel(main_filename, sheet_name = 'To Transfer', skiprows = 4)
    marker = len(df_2.index)
    df_2 = df_2.append(df_3)
    
    wb = opx.load_workbook(filename)
    ws = wb['Sheet1']
    
    
    sum_due = 0
    sum_mon = 0
    sum_bil = 0
    
    for rows in range(3, 15):
        for index_2 in range (0,len(df_2.index)):
            if not isinstance(df_2.iloc[index_2]['RI PIC'], float):
                if ri_pic in df_2.iloc[index_2]['RI PIC']:
                    if df_2.iloc[index_2]['Customer'].upper() == "SIGN": 
                        customer = df_2.iloc[index_2]['Customer'].upper()
                    else: customer = df_2.iloc[index_2]['Customer']
                    for index_1 in range (0,len(df_1.index)):
                        if customer in df_1.iloc[index_1]['Customer Name']:
                            if index_2<= marker:
                                if datetime.strftime(df_2.iloc[index_2]['Month Of Transfer'],'%b') in calendar.month_name[(rows)%12+1]:
                                    sum_bil += df_2.iloc[index_2]['Amount charged to customer (without GST)']
                                    sum_mon += df_2.iloc[index_2]['Amount to transfer']
                            for columns in range(1, ws.max_column+1):
                                if str(df_2.iloc[index_2]['Invoice Date'])[3:5].isdigit():
                                    if calendar.month_name[int(df_2.iloc[index_2]['Invoice Date'][3:5])] in ws.cell(rows,1).value:
                                        if ws.cell(1,columns).value in df_1.iloc[index_1]['Category']:
                                            sum_due += df_2.iloc[index_2]['Amount to transfer']
                                            if ws.cell(rows,columns).value is not None:
                                                ws.cell(rows,columns).value += df_2.iloc[index_2]['Amount to transfer']
                                            else: ws.cell(rows,columns).value = df_2.iloc[index_2]['Amount to transfer']
        ws.cell(rows,2).value = sum_bil
        ws.cell(rows,ws.max_column-2).value = sum_mon
        sum_bil = sum_mon = 0
        wb.save(filename)
                                        
    ###################################################################################################################
    #Balance carry forward from FY18
    ws.cell(2,ws.max_column).value = 0
    
    wb.save(filename)
    wb.close()
    
def main():
    
    #will check for all files from script directory onwards
    print (os.getcwd(), ":") #get dir address for current file
    files_dir = list()
    for root, dirs, files in os.walk('.'):
        for file in files:
            if '.xlsx' in file or '.csv' in file:
                files_dir.append(file)
                print ("\t", file)
    print ("")
    
    temp_created = False
    custlist_found = False
    for filename in files_dir:
        if filename[-4:]=='.csv': 
            filename = convert_to_xlsx(filename)
            temp_created = True
        clean(filename)
        
        if 'Customer List' in filename: custlist_found = True
        if 'Transfer of Revenue to RIs' in filename: extract_pis(filename)
        if temp_created == True: os.remove(filename)
        
    if custlist_found == False: print ("Customer List not found")
if __name__ == '__main__':
    main()