# -*- coding: utf-8 -*-
"""
Created on Wed May  8 15:05:12 2019

@author: tp-AizatKA
"""

import os
import sys
import math
import pandas as pd
import openpyxl as opx
from openpyxl.styles import PatternFill

#pip install --upgrade easygui
from easygui import choicebox

from datetime import datetime

global DDMM

#extract out booking hours = 0 and fee/hour not NA
def extract_ext(file_raw, temp_file, zero_key, nonzero_key):
    
    #open excel file as df by pandas
    df = pd.read_excel(file_raw)
    writer = pd.ExcelWriter(temp_file)
    
    to_include = list()
    
    for index_i, i in enumerate(df.filter([zero_key])[zero_key]):
        if i == 0 and not math.isnan(df.filter([nonzero_key])[nonzero_key][index_i]): to_include.append(index_i)

    #write to file
    df.loc[to_include].to_excel(writer)
        
    writer.save()
    writer.close()
    
#read and apply function to the specific column
def adj_amount(temp_file):
    
    wb = opx.load_workbook(temp_file)
    ws = wb['Sheet1']

    ws['Q1'] = "Adjusted Duration Used (in Hours)"
    ws['R1'] = "Amount to Bill\n(Column M x Column P x (1-Subsidy)"
    
    prev_col = "text variable"
    for column, column_names in enumerate(ws['1'], 1):
        #all other columns referenceed off this column
        if prev_col == 'Effective Cost': #by title of column Q
            for row in range(2, ws.max_row+1):
                ws.cell(row, column).value = round_uptom_max(ws.cell(row, column-4).value, 0.5)
                #Fee/Hour                  -8
                #Fee for this Use          -3
                #Subsidy (%)               -2 
                if ws.cell(row, column-8).value is None: rate = 0
                else: rate = ws.cell(row, column-8).value
                subsidy = ws.cell(row, column-2).value
                ws.cell(row, column+1).value = ws.cell(row, column).value * rate * (1 - subsidy)
        prev_col = column_names.value
    wb.save(temp_file)

#round up value to nearest multiple of m with max value of 2.0    
def round_uptom_max(n, m):
    
    n *= 10
    m *= 10
    if n%m != 0: n = ((n//m)+1)*m
    if n > 20: n = 20
    return n/10

#compile similar sessions to create an entry type as of billing document
def consolidate(temp_file):
    
    df = pd.read_excel(temp_file)
    df2 = []
    del_list = []
    total_hrs = total_amt = count = 0

    key = 'User Full Name (Last-First)'
    df.sort_values([key], inplace = True)
    total_hrs += df.iloc[0][15]
    total_amt += df.iloc[0][16]
    count += 1
    for i in range(1, len(df.index)): 
        if df.iloc[i-1][3] == df.iloc[i][3]: 
            total_hrs += df.iloc[i][15]
            total_amt += df.iloc[i][16]
            count += 1
            del_list.append(i)
        else:
            #In order of: Billing name, PI name, User name, NUmber of sessions, Used hours, Total Amount
            df2.append([df.iloc[i-1, 1], df.iloc[i-1, 2], df.iloc[i-1, 3], count, total_hrs, total_amt])
            total_hrs = df.iloc[i][15]
            total_amt = df.iloc[i][16]
            count = 1
    df2.append([df.iloc[i-1, 1], df.iloc[i-1, 2], df.iloc[i-1, 3], count, total_hrs, total_amt])
    return df2
    
#append and copy other data where relevant
def append_to_report(report_file, report_file_sheet, df2):
    
    wb = opx.load_workbook(report_file)
    ws = wb[report_file_sheet]
    
    point = ws.max_row
    for index_entry, entry in enumerate(df2, 1):
        i = point+index_entry
        for j in range(1,ws.max_column+1):
            if ws.cell(1, j).value == 'Billing Name (non-A*STAR)/Research Institute (A*STAR)': 
                ws.cell(i, j).value=entry[0]
            if ws.cell(1, j).value == 'PI name': 
                ws.cell(i, j).value=entry[1]
            if ws.cell(1, j).value == 'User name': 
                ws.cell(i, j).value=entry[2]
            if ws.cell(1, j).value == 'SAP material code': 
                ws.cell(i, j).value=ws.cell(2, j).value
                #ws.cell(i, j).value='SV-BD-0042'
            if ws.cell(1, j).value == 'Description': 
                ws.cell(i, j).value=ws.cell(2, j).value
            if ws.cell(1, j).value == 'NUmber of sessions': 
                ws.cell(i, j).value=entry[3]
            if ws.cell(1, j).value == 'Booked hours': 
                ws.cell(i, j).value=0
            if ws.cell(1, j).value == 'Used hours': 
                ws.cell(i, j).value=entry[4]
            if ws.cell(1, j).value == 'Amount to bill': 
                ws.cell(i, j).value=entry[5]
            if ws.cell(1, j).value == 'Total amount':
                if ws.cell(i, j-1).value is None:
                    ws.cell(i, j).value=entry[5]
                else: ws.cell(i, j).value=ws.cell(i, j-1).value+ws.cell(i, j-2).value
            if ws.cell(1, j).value == 'Group manager accepted':
                ws.cell(i, j).value='Accepted by Default'
                
    #blue fill
    for row in range(point+1,ws.max_row+1):
        for col in range(1,ws.max_column+1):
            ws.cell(row, col).fill = PatternFill(fgColor = '99CCFF', fill_type = 'solid')
            
    #get other columns data
    condition = []
    for new_set in range(point+1,ws.max_row+1):
        for old_set in range(1,point+1):
            for col_to_check in [3,6,7]:
                #check if same in column C, F, G
                condition.append(ws.cell(new_set, col_to_check).value == ws.cell(old_set, col_to_check).value)
            if False not in condition: 
                for col_to_change in [1,2,4,5,8,9,10]:
                    ws.cell(new_set, col_to_change).value = ws.cell(old_set, col_to_change).value
                    if '(EXTERNAL)' in ws.cell(new_set, 1).value.upper():
                        ws.cell(new_set, 20).value=(ws.cell(new_set, 19).value*0.07)
            condition = []
        
    DDMM = datetime.now().strftime('%d') + datetime.now().strftime('%m')
    
    final_file = report_file.replace('.xlsx','') + '_py' + DDMM + '.xlsx'
    wb.save(final_file)
    return final_file

#converts .csv files to .xlsx files for compatibality
def convert_to_xlsx(init_filename):
    import csv
    from openpyxl import Workbook

    
    f = open(init_filename)
    csv.register_dialect('comma', delimiter=',')
    reader = csv.reader(f, dialect='comma')
    
    wb = Workbook()
    dest_filename = init_filename.replace('.csv', '.xlsx')
    
    ws = wb.worksheets[0]
    ws.title = "Sheet 1"
    
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):

            ws.cell(row_index + 1, column_index + 1).value = cell
    
    wb.save(filename = dest_filename)
    
    clean(dest_filename)
    
    return dest_filename

#clean whole file from special character '|'
def clean(file):
    wb = opx.load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in range(1, ws.max_column):
            for row in range(1, ws.max_row + 1):
                if type(ws.cell(row, col).value) is str:
                    # .replace('changefrom', 'changeto') and .strip() removes leading/trailing spaces
                    ws.cell(row, col).value = ws.cell(row, col).value.replace('|', '').strip()
                    ws.cell(row, col).value = ws.cell(row, col).value.replace('ï»¿', '').strip()
                    if ws.cell(row, col).value is not None and ws.cell(row, col).value.isdigit(): ws.cell(row, col).value = int(ws.cell(row, col).value)
                #assign 0 for blanks under 'Bill to....' column
                if col == 2 and ws.cell(row, col).value is None: ws.cell(row, col).value = 0
    wb.save(file)
    
def main():

    #will check for all files from script directory onwards
    print (os.getcwd(), ":") #get dir address for current file
    files_dir = list()
    for root, dirs, files in os.walk('.'):
        for file in files:
            if '.xlsx' in file or '.csv' in file:
                files_dir.append(file)
                print ("\t", file)
    print ("")
    
    #dialog box to choose file
    msg ="Please select the raw generated file\nor Press <Cancel> to exit."
    title = "Directory List of .xlsx Files"
    choice1 = choicebox(msg, title, files_dir)
    print ("Raw file selected:", choice1)
    
    msg ="Please select the BILLING REPORT\nor Press <Cancel> to exit."
    title = "Directory List of .xlsx Files"
    choice2 = choicebox(msg, title, files_dir)
    print ("Billing report selected:", choice2)
    
    if choice1 is None or choice2 is None:
        sys.exit("Files not selected..")

    temp_created = False
    for filename in [choice1, choice2]:
        if filename[-4:]=='.csv': 
            filename = convert_to_xlsx(filename)
            temp_created = True
            
    file_raw = choice1.replace('.csv','.xlsx')
    temp_file = 'new.xlsx'
    zero_key = 'Duration Booked (in Hours)'
    nonzero_key = 'Fee/Hour'
    report_file = choice2.replace('.csv','.xlsx')
    report_file_sheet = 'export'
    
    extract_ext(file_raw, temp_file, zero_key, nonzero_key)
    adj_amount(temp_file)
    to_add = consolidate(temp_file)
    final_file = append_to_report(report_file, report_file_sheet, to_add)
    
    print (final_file, "has been created!")
    
    os.remove(temp_file)
    if temp_created == True:
        if file_raw != choice1:
            os.remove(file_raw)
        if report_file != choice2:
            os.remove(report_file)
    
if __name__ == '__main__':
    main()