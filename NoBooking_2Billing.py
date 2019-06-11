# -*- coding: utf-8 -*-
"""
Created on Thu May 30 17:04:32 2019

@author: tp-AizatKA
"""
    
import os
import pandas as pd
import openpyxl as opx

#pip install --upgrade easygui
from easygui import choicebox

#update gst
def update_gst(final_file, report_file_sheet):
    wb = opx.load_workbook(final_file)
    ws = wb[report_file_sheet]
    for col in range(1, ws.max_column+1):
        if ws.cell(1, col).value == 'GST':
            for row in range(1, ws.max_row+1):
                if ws.cell(row, 1).value is not None:
                    if '(EXTERNAL)' in ws.cell(row, 1).value.upper() and ws.cell(row, col).value is None:
                        ws.cell(row, col).value = ws.cell(row, col-1).value*0.07
    wb.save(final_file)

#combine similar data entry of the generated and calculated for billing
def concise(final_file, report_file_sheet):
    
    df = pd.read_excel(final_file)
    
    df.sort_values(['User name', 'NUmber of sessions'], inplace = True)
    df.reset_index(drop=True, inplace=True)

    point = len(df.index)
    del_list = []
    for i in range(1, point): 
        if df.iloc[i-1]['User name'] == df.iloc[i]['User name']:
            if not df.iloc[i-1]['NUmber of sessions']==df.iloc[i-1]['Booked hours']==df.iloc[i-1]['Used hours']==0:
                if not df.iloc[i]['NUmber of sessions']==df.iloc[i]['Booked hours']==df.iloc[i]['Used hours']==0:
                    del_list.append(i)
                    for j in [13,14,15,18,20]: df.iloc[i-1,j]+=df.iloc[i,j]
                    
                    #GST column; due to manipulation of NaN
                    if df.isnull().iloc[i-1,19]: df.iloc[i-1,19] = df.iloc[i,19]
                    elif df.isnull().iloc[i,19]: df.iloc[i,19] = df.iloc[i-1,19]
                    else: df.iloc[i-1,19]+=df.iloc[i,19]
      
    #delete consolidated data                
    df.drop(del_list, inplace=True)
    
    df.sort_values(['Billing Name (non-A*STAR)/Research Institute (A*STAR)', 'Account number'], inplace = True)
    
    #write to file
    concised_file = final_file.replace('.xlsx','')+'_2.xlsx'
    writer = pd.ExcelWriter(concised_file)
    df.to_excel(writer, index=False)
    writer.save()
    writer.close()
    return concised_file

def main():
    
    #will check for all files from script directory onwards
    print (os.getcwd(), ":") #get dir address for current file
    files_dir = list()
    for root, dirs, files in os.walk('.'):
        for file in files:
            if '.xlsx' in file:
                files_dir.append(file)
                print ("\t", file)
    print ("")
    
    report_file_sheet = 'export'
    
    #dialog box to choose file
    msg ="Please choose file to CONSOLIDATE\nor Press <Cancel> to exit."
    title = "Directory List of .xlsx Files"
    final_file = choicebox(msg, title, files_dir)
    
    update_gst(final_file, report_file_sheet)
    concised_file = concise(final_file, report_file_sheet)
    
    print (concised_file, "has been created!")
    
if __name__ == '__main__':
    main()